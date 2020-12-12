import openpyxl

book = openpyxl.load_workbook("E:\\Tech bite\\4_Selenium_Udemy\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
# or
print(sheet['B1'].value)

sheet.cell(row=2, column=2).value = "Ramesh"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row)
print(sheet.max_column)



